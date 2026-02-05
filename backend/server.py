from fastapi import FastAPI, APIRouter, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
from datetime import datetime, date
from bson import ObjectId
import io
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
from starlette.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def serialize_doc(doc):
    if doc and '_id' in doc:
        doc['id'] = str(doc['_id'])
        del doc['_id']
    return doc

# ============= MODELS =============

class PartSpec(BaseModel):
    part_name: str
    quantity_needed: float

class Item(BaseModel):
    name: str
    category: str  # "Street Light" or "Flood Light"
    parts: List[PartSpec]
    opening_stock: float = 0
    current_stock: float = 0
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ItemResponse(BaseModel):
    id: str
    name: str
    category: str
    parts: List[Dict]
    opening_stock: float
    current_stock: float
    created_at: datetime

class PartStock(BaseModel):
    part_name: str
    opening_stock: float = 0
    current_stock: float = 0
    created_at: datetime = Field(default_factory=datetime.utcnow)

class PartStockResponse(BaseModel):
    id: str
    part_name: str
    opening_stock: float
    current_stock: float
    created_at: datetime

class ProductionEntry(BaseModel):
    date: str
    item_id: str
    item_name: str
    quantity: int

class ProductionResponse(BaseModel):
    id: str
    date: str
    item_id: str
    item_name: str
    quantity: int
    created_at: datetime

class SalesEntry(BaseModel):
    date: str
    item_id: str
    item_name: str
    quantity: int
    party_name: str

class SalesResponse(BaseModel):
    id: str
    date: str
    item_id: str
    item_name: str
    quantity: int
    party_name: str
    created_at: datetime

class PurchaseEntry(BaseModel):
    date: str
    item_id: str
    part_name: str
    quantity: float

class PurchaseResponse(BaseModel):
    id: str
    date: str
    item_id: str
    item_name: str
    part_name: str
    quantity: float
    created_at: datetime

# ============= ITEMS API =============

@api_router.post("/items", response_model=ItemResponse)
async def create_item(item: Item):
    try:
        item_dict = item.dict()
        item_dict['current_stock'] = item.opening_stock
        
        result = await db.items.insert_one(item_dict)
        
        # Create part stock entries for each part
        for part_spec in item.parts:
            existing_part = await db.part_stocks.find_one({"part_name": part_spec.part_name})
            if not existing_part:
                part_stock = PartStock(part_name=part_spec.part_name)
                await db.part_stocks.insert_one(part_stock.dict())
        
        created = await db.items.find_one({"_id": result.inserted_id})
        return serialize_doc(created)
    except Exception as e:
        logger.error(f"Error creating item: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/items", response_model=List[ItemResponse])
async def get_items():
    try:
        items = await db.items.find().sort("name", 1).to_list(1000)
        return [serialize_doc(item) for item in items]
    except Exception as e:
        logger.error(f"Error fetching items: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/items/{item_id}", response_model=ItemResponse)
async def get_item(item_id: str):
    try:
        item = await db.items.find_one({"_id": ObjectId(item_id)})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        return serialize_doc(item)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching item: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/items/{item_id}", response_model=ItemResponse)
async def update_item(item_id: str, item: Item):
    try:
        # Update the item
        item_dict = item.dict()
        result = await db.items.update_one(
            {"_id": ObjectId(item_id)},
            {"$set": item_dict}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Create part stock entries for new parts
        for part_spec in item.parts:
            existing_part = await db.part_stocks.find_one({"part_name": part_spec.part_name})
            if not existing_part:
                part_stock = PartStock(part_name=part_spec.part_name)
                await db.part_stocks.insert_one(part_stock.dict())
        
        updated = await db.items.find_one({"_id": ObjectId(item_id)})
        return serialize_doc(updated)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating item: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============= PART STOCKS API =============

@api_router.get("/part-stocks", response_model=List[PartStockResponse])
async def get_part_stocks():
    try:
        parts = await db.part_stocks.find().sort("part_name", 1).to_list(1000)
        return [serialize_doc(part) for part in parts]
    except Exception as e:
        logger.error(f"Error fetching part stocks: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/part-stocks", response_model=PartStockResponse)
async def create_part_stock(part: PartStock):
    try:
        # Check if part already exists
        existing = await db.part_stocks.find_one({
            "part_name": part.part_name
        })

        # If exists → UPDATE
        if existing:
            new_stock = existing["current_stock"] + part.quantity

            await db.part_stocks.update_one(
                {"_id": existing["_id"]},
                {
                    "$set": {
                        "current_stock": new_stock
                    }
                }
            )

            updated = await db.part_stocks.find_one(
                {"_id": existing["_id"]}
            )

            return serialize_doc(updated)

        # If not exists → CREATE
        part_dict = part.dict()
        part_dict["opening_stock"] = part.quantity
        part_dict["current_stock"] = part.quantity

        result = await db.part_stocks.insert_one(part_dict)

        created = await db.part_stocks.find_one(
            {"_id": result.inserted_id}
        )

        return serialize_doc(created)

    except Exception as e:
        logger.error(f"Error creating part stock: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============= PRODUCTION API =============

@api_router.post("/production", response_model=ProductionResponse)
async def create_production(production: ProductionEntry):
    try:
        # Get item
        item = await db.items.find_one({"_id": ObjectId(production.item_id)})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Check if parts are available
        insufficient_parts = []
        for part_spec in item['parts']:
            # Find part with highest stock (to handle duplicates)
            part_stocks = await db.part_stocks.find({"part_name": part_spec['part_name']}).sort("current_stock", -1).to_list(10)
            if not part_stocks:
                insufficient_parts.append({
                    "part_name": part_spec['part_name'],
                    "required": part_spec['quantity_needed'] * production.quantity,
                    "available": 0
                })
            else:
                part_stock = part_stocks[0]  # Get the one with highest stock
                if part_stock['current_stock'] < (part_spec['quantity_needed'] * production.quantity):
                    insufficient_parts.append({
                        "part_name": part_spec['part_name'],
                        "required": part_spec['quantity_needed'] * production.quantity,
                        "available": part_stock['current_stock']
                    })
        
        if insufficient_parts:
            raise HTTPException(
                status_code=400,
                detail={"message": "Insufficient parts", "parts": insufficient_parts}
            )
        
        # Deduct parts
        for part_spec in item['parts']:
            # Find part with highest stock (to handle duplicates)
            part_stocks = await db.part_stocks.find({"part_name": part_spec['part_name']}).sort("current_stock", -1).to_list(10)
            part_stock = part_stocks[0]  # Get the one with highest stock
            new_stock = part_stock['current_stock'] - (part_spec['quantity_needed'] * production.quantity)
            await db.part_stocks.update_one(
                {"_id": part_stock['_id']},
                {"$set": {"current_stock": new_stock}}
            )
        
        # Add finished goods
        new_item_stock = item['current_stock'] + production.quantity
        await db.items.update_one(
            {"_id": ObjectId(production.item_id)},
            {"$set": {"current_stock": new_item_stock}}
        )
        
        # Record production
        production_dict = production.dict()
        production_dict['created_at'] = datetime.utcnow()
        result = await db.production.insert_one(production_dict)
        created = await db.production.find_one({"_id": result.inserted_id})
        
        return serialize_doc(created)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating production: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/production", response_model=List[ProductionResponse])
async def get_production():
    try:
        production = await db.production.find().sort("date", -1).to_list(1000)
        return [serialize_doc(p) for p in production]
    except Exception as e:
        logger.error(f"Error fetching production: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/production/{production_id}")
async def delete_production(production_id: str):
    try:
        # Get production record
        production = await db.production.find_one({"_id": ObjectId(production_id)})
        if not production:
            raise HTTPException(status_code=404, detail="Production record not found")
        
        # Get item
        item = await db.items.find_one({"_id": ObjectId(production['item_id'])})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Reverse: Add parts back
        for part_spec in item['parts']:
            part_stock = await db.part_stocks.find_one({"part_name": part_spec['part_name']})
            if part_stock:
                new_stock = part_stock['current_stock'] + (part_spec['quantity_needed'] * production['quantity'])
                await db.part_stocks.update_one(
                    {"_id": part_stock['_id']},
                    {"$set": {"current_stock": new_stock}}
                )
        
        # Reverse: Remove finished goods
        new_item_stock = item['current_stock'] - production['quantity']
        await db.items.update_one(
            {"_id": ObjectId(production['item_id'])},
            {"$set": {"current_stock": new_item_stock}}
        )
        
        # Delete production record
        await db.production.delete_one({"_id": ObjectId(production_id)})
        
        return {"message": "Production deleted successfully", "reversed_quantity": production['quantity']}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting production: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/production/{production_id}")
async def update_production(production_id: str, new_quantity: int):
    try:
        # Get production record
        production = await db.production.find_one({"_id": ObjectId(production_id)})
        if not production:
            raise HTTPException(status_code=404, detail="Production record not found")
        
        old_quantity = production['quantity']
        quantity_diff = new_quantity - old_quantity
        
        if quantity_diff == 0:
            return {"message": "No change in quantity"}
        
        # Get item
        item = await db.items.find_one({"_id": ObjectId(production['item_id'])})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Check if we have enough parts for increase
        if quantity_diff > 0:
            insufficient_parts = []
            for part_spec in item['parts']:
                part_stock = await db.part_stocks.find_one({"part_name": part_spec['part_name']})
                if not part_stock:
                    insufficient_parts.append({
                        "part_name": part_spec['part_name'],
                        "required": part_spec['quantity_needed'] * quantity_diff,
                        "available": 0
                    })
                elif part_stock['current_stock'] < (part_spec['quantity_needed'] * quantity_diff):
                    insufficient_parts.append({
                        "part_name": part_spec['part_name'],
                        "required": part_spec['quantity_needed'] * quantity_diff,
                        "available": part_stock['current_stock']
                    })
            
            if insufficient_parts:
                raise HTTPException(
                    status_code=400,
                    detail={"message": "Insufficient parts for increase", "parts": insufficient_parts}
                )
        
        # Update parts stock
        for part_spec in item['parts']:
            part_stock = await db.part_stocks.find_one({"part_name": part_spec['part_name']})
            if part_stock:
                # If increasing production, deduct more parts
                # If decreasing production, add parts back
                new_stock = part_stock['current_stock'] - (part_spec['quantity_needed'] * quantity_diff)
                await db.part_stocks.update_one(
                    {"_id": part_stock['_id']},
                    {"$set": {"current_stock": new_stock}}
                )
        
        # Update finished goods
        new_item_stock = item['current_stock'] + quantity_diff
        await db.items.update_one(
            {"_id": ObjectId(production['item_id'])},
            {"$set": {"current_stock": new_item_stock}}
        )
        
        # Update production record
        await db.production.update_one(
            {"_id": ObjectId(production_id)},
            {"$set": {"quantity": new_quantity}}
        )
        
        return {
            "message": "Production updated successfully",
            "old_quantity": old_quantity,
            "new_quantity": new_quantity,
            "change": quantity_diff
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating production: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/production/{production_id}", response_model=ProductionResponse)
async def update_production(production_id: str, production: ProductionEntry):
    try:
        # Get existing production entry
        existing = await db.production.find_one({"_id": ObjectId(production_id)})
        if not existing:
            raise HTTPException(status_code=404, detail="Production entry not found")
        
        # Get item
        item = await db.items.find_one({"_id": ObjectId(production.item_id)})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Reverse the old production (add parts back, reduce finished goods)
        old_item = await db.items.find_one({"_id": ObjectId(existing['item_id'])})
        if old_item:
            # Add parts back
            for part_spec in old_item['parts']:
                part_stocks = await db.part_stocks.find({"part_name": part_spec['part_name']}).sort("current_stock", 1).to_list(10)
                if part_stocks:
                    part_stock = part_stocks[0]  # Get the one with lowest stock to balance
                    new_stock = part_stock['current_stock'] + (part_spec['quantity_needed'] * existing['quantity'])
                    await db.part_stocks.update_one(
                        {"_id": part_stock['_id']},
                        {"$set": {"current_stock": new_stock}}
                    )
            
            # Reduce finished goods
            new_item_stock = old_item['current_stock'] - existing['quantity']
            await db.items.update_one(
                {"_id": ObjectId(existing['item_id'])},
                {"$set": {"current_stock": new_item_stock}}
            )
        
        # Check if parts are available for new production
        insufficient_parts = []
        for part_spec in item['parts']:
            part_stocks = await db.part_stocks.find({"part_name": part_spec['part_name']}).sort("current_stock", -1).to_list(10)
            if not part_stocks:
                insufficient_parts.append({
                    "part_name": part_spec['part_name'],
                    "required": part_spec['quantity_needed'] * production.quantity,
                    "available": 0
                })
            else:
                part_stock = part_stocks[0]
                if part_stock['current_stock'] < (part_spec['quantity_needed'] * production.quantity):
                    insufficient_parts.append({
                        "part_name": part_spec['part_name'],
                        "required": part_spec['quantity_needed'] * production.quantity,
                        "available": part_stock['current_stock']
                    })
        
        if insufficient_parts:
            raise HTTPException(
                status_code=400,
                detail={"message": "Insufficient parts", "parts": insufficient_parts}
            )
        
        # Apply new production (deduct parts, add finished goods)
        for part_spec in item['parts']:
            part_stocks = await db.part_stocks.find({"part_name": part_spec['part_name']}).sort("current_stock", -1).to_list(10)
            part_stock = part_stocks[0]
            new_stock = part_stock['current_stock'] - (part_spec['quantity_needed'] * production.quantity)
            await db.part_stocks.update_one(
                {"_id": part_stock['_id']},
                {"$set": {"current_stock": new_stock}}
            )
        
        # Add finished goods
        current_item = await db.items.find_one({"_id": ObjectId(production.item_id)})
        new_item_stock = current_item['current_stock'] + production.quantity
        await db.items.update_one(
            {"_id": ObjectId(production.item_id)},
            {"$set": {"current_stock": new_item_stock}}
        )
        
        # Update production record
        production_dict = production.dict()
        await db.production.update_one(
            {"_id": ObjectId(production_id)},
            {"$set": production_dict}
        )
        
        updated = await db.production.find_one({"_id": ObjectId(production_id)})
        return serialize_doc(updated)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating production: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/production/{production_id}")
async def delete_production(production_id: str):
    try:
        # Get existing production entry
        existing = await db.production.find_one({"_id": ObjectId(production_id)})
        if not existing:
            raise HTTPException(status_code=404, detail="Production entry not found")
        
        # Get item
        item = await db.items.find_one({"_id": ObjectId(existing['item_id'])})
        if item:
            # Add parts back
            for part_spec in item['parts']:
                part_stocks = await db.part_stocks.find({"part_name": part_spec['part_name']}).sort("current_stock", 1).to_list(10)
                if part_stocks:
                    part_stock = part_stocks[0]  # Get the one with lowest stock to balance
                    new_stock = part_stock['current_stock'] + (part_spec['quantity_needed'] * existing['quantity'])
                    await db.part_stocks.update_one(
                        {"_id": part_stock['_id']},
                        {"$set": {"current_stock": new_stock}}
                    )
            
            # Reduce finished goods
            new_item_stock = item['current_stock'] - existing['quantity']
            await db.items.update_one(
                {"_id": ObjectId(existing['item_id'])},
                {"$set": {"current_stock": new_item_stock}}
            )
        
        # Delete production record
        await db.production.delete_one({"_id": ObjectId(production_id)})
        
        return {"message": "Production entry deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting production: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============= SALES API =============

@api_router.post("/sales", response_model=SalesResponse)
async def create_sale(sale: SalesEntry):
    try:
        # Get item
        item = await db.items.find_one({"_id": ObjectId(sale.item_id)})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Check stock
        if item['current_stock'] < sale.quantity:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient stock. Available: {item['current_stock']}, Requested: {sale.quantity}"
            )
        
        # Reduce stock
        new_stock = item['current_stock'] - sale.quantity
        await db.items.update_one(
            {"_id": ObjectId(sale.item_id)},
            {"$set": {"current_stock": new_stock}}
        )
        
        # Record sale
        sale_dict = sale.dict()
        sale_dict['created_at'] = datetime.utcnow()
        result = await db.sales.insert_one(sale_dict)
        created = await db.sales.find_one({"_id": result.inserted_id})
        
        return serialize_doc(created)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating sale: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/sales", response_model=List[SalesResponse])
async def get_sales():
    try:
        sales = await db.sales.find().sort("date", -1).to_list(1000)
        return [serialize_doc(s) for s in sales]
    except Exception as e:
        logger.error(f"Error fetching sales: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/sales/{sale_id}")
async def delete_sale(sale_id: str):
    try:
        # Get sale record
        sale = await db.sales.find_one({"_id": ObjectId(sale_id)})
        if not sale:
            raise HTTPException(status_code=404, detail="Sale record not found")
        
        # Get item
        item = await db.items.find_one({"_id": ObjectId(sale['item_id'])})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Reverse: Add finished goods back
        new_stock = item['current_stock'] + sale['quantity']
        await db.items.update_one(
            {"_id": ObjectId(sale['item_id'])},
            {"$set": {"current_stock": new_stock}}
        )
        
        # Delete sale record
        await db.sales.delete_one({"_id": ObjectId(sale_id)})
        
        return {"message": "Sale deleted successfully", "reversed_quantity": sale['quantity']}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting sale: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/sales/{sale_id}")
async def update_sale(sale_id: str, new_quantity: int, party_name: str = None):
    try:
        # Get sale record
        sale = await db.sales.find_one({"_id": ObjectId(sale_id)})
        if not sale:
            raise HTTPException(status_code=404, detail="Sale record not found")
        
        old_quantity = sale['quantity']
        quantity_diff = new_quantity - old_quantity
        
        # Get item
        item = await db.items.find_one({"_id": ObjectId(sale['item_id'])})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Check if we have enough stock for increase
        if quantity_diff > 0:
            if item['current_stock'] < quantity_diff:
                raise HTTPException(
                    status_code=400,
                    detail=f"Insufficient stock. Available: {item['current_stock']}, Additional needed: {quantity_diff}"
                )
        
        # Update finished goods stock
        new_stock = item['current_stock'] - quantity_diff
        await db.items.update_one(
            {"_id": ObjectId(sale['item_id'])},
            {"$set": {"current_stock": new_stock}}
        )
        
        # Update sale record
        update_fields = {"quantity": new_quantity}
        if party_name:
            update_fields["party_name"] = party_name
            
        await db.sales.update_one(
            {"_id": ObjectId(sale_id)},
            {"$set": update_fields}
        )
        
        return {
            "message": "Sale updated successfully",
            "old_quantity": old_quantity,
            "new_quantity": new_quantity,
            "change": quantity_diff
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating sale: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/sales/{sale_id}", response_model=SalesResponse)
async def update_sale(sale_id: str, sale: SalesEntry):
    try:
        # Get existing sale entry
        existing = await db.sales.find_one({"_id": ObjectId(sale_id)})
        if not existing:
            raise HTTPException(status_code=404, detail="Sale entry not found")
        
        # Get item
        item = await db.items.find_one({"_id": ObjectId(sale.item_id)})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Reverse the old sale (add stock back)
        old_item = await db.items.find_one({"_id": ObjectId(existing['item_id'])})
        if old_item:
            new_stock = old_item['current_stock'] + existing['quantity']
            await db.items.update_one(
                {"_id": ObjectId(existing['item_id'])},
                {"$set": {"current_stock": new_stock}}
            )
        
        # Check stock for new sale
        current_item = await db.items.find_one({"_id": ObjectId(sale.item_id)})
        if current_item['current_stock'] < sale.quantity:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient stock. Available: {current_item['current_stock']}, Requested: {sale.quantity}"
            )
        
        # Apply new sale (reduce stock)
        new_stock = current_item['current_stock'] - sale.quantity
        await db.items.update_one(
            {"_id": ObjectId(sale.item_id)},
            {"$set": {"current_stock": new_stock}}
        )
        
        # Update sale record
        sale_dict = sale.dict()
        await db.sales.update_one(
            {"_id": ObjectId(sale_id)},
            {"$set": sale_dict}
        )
        
        updated = await db.sales.find_one({"_id": ObjectId(sale_id)})
        return serialize_doc(updated)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating sale: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/sales/{sale_id}")
async def delete_sale(sale_id: str):
    try:
        # Get existing sale entry
        existing = await db.sales.find_one({"_id": ObjectId(sale_id)})
        if not existing:
            raise HTTPException(status_code=404, detail="Sale entry not found")
        
        # Get item and add stock back
        item = await db.items.find_one({"_id": ObjectId(existing['item_id'])})
        if item:
            new_stock = item['current_stock'] + existing['quantity']
            await db.items.update_one(
                {"_id": ObjectId(existing['item_id'])},
                {"$set": {"current_stock": new_stock}}
            )
        
        # Delete sale record
        await db.sales.delete_one({"_id": ObjectId(sale_id)})
        
        return {"message": "Sale entry deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting sale: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============= PURCHASE API =============

@api_router.post("/purchases", response_model=PurchaseResponse)
async def create_purchase(purchase: PurchaseEntry):
    try:
        # Get item for reference
        item = await db.items.find_one({"_id": ObjectId(purchase.item_id)})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Update part stock
        part_stock = await db.part_stocks.find_one({"part_name": purchase.part_name})
        if not part_stock:
            # Create if doesn't exist
            part_stock = PartStock(part_name=purchase.part_name, opening_stock=0, current_stock=0)
            result = await db.part_stocks.insert_one(part_stock.dict())
            part_stock = await db.part_stocks.find_one({"_id": result.inserted_id})
        
        new_stock = part_stock['current_stock'] + purchase.quantity
        await db.part_stocks.update_one(
            {"_id": part_stock['_id']},
            {"$set": {"current_stock": new_stock}}
        )
        
        # Record purchase
        purchase_dict = purchase.dict()
        purchase_dict['item_name'] = item['name']
        purchase_dict['created_at'] = datetime.utcnow()
        result = await db.purchases.insert_one(purchase_dict)
        created = await db.purchases.find_one({"_id": result.inserted_id})
        
        return serialize_doc(created)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating purchase: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/purchases", response_model=List[PurchaseResponse])
async def get_purchases():
    try:
        purchases = await db.purchases.find().sort("date", -1).to_list(1000)
        return [serialize_doc(p) for p in purchases]
    except Exception as e:
        logger.error(f"Error fetching purchases: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str):
    try:
        # Get purchase record
        purchase = await db.purchases.find_one({"_id": ObjectId(purchase_id)})
        if not purchase:
            raise HTTPException(status_code=404, detail="Purchase record not found")
        
        # Update part stock (reverse)
        part_stock = await db.part_stocks.find_one({"part_name": purchase['part_name']})
        if part_stock:
            new_stock = part_stock['current_stock'] - purchase['quantity']
            if new_stock < 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot delete: would result in negative stock for {purchase['part_name']}"
                )
            await db.part_stocks.update_one(
                {"_id": part_stock['_id']},
                {"$set": {"current_stock": new_stock}}
            )
        
        # Delete purchase record
        await db.purchases.delete_one({"_id": ObjectId(purchase_id)})
        
        return {"message": "Purchase deleted successfully", "reversed_quantity": purchase['quantity']}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting purchase: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/purchases/{purchase_id}", response_model=PurchaseResponse)
async def update_purchase(purchase_id: str, purchase: PurchaseEntry):
    try:
        # Get existing purchase entry
        existing = await db.purchases.find_one({"_id": ObjectId(purchase_id)})
        if not existing:
            raise HTTPException(status_code=404, detail="Purchase entry not found")
        
        # Get item for reference
        item = await db.items.find_one({"_id": ObjectId(purchase.item_id)})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Reverse the old purchase (reduce part stock)
        old_part_stock = await db.part_stocks.find_one({"part_name": existing['part_name']})
        if old_part_stock:
            new_stock = old_part_stock['current_stock'] - existing['quantity']
            await db.part_stocks.update_one(
                {"_id": old_part_stock['_id']},
                {"$set": {"current_stock": new_stock}}
            )
        
        # Apply new purchase (add part stock)
        part_stock = await db.part_stocks.find_one({"part_name": purchase.part_name})
        if not part_stock:
            # Create if doesn't exist
            part_stock = PartStock(part_name=purchase.part_name, opening_stock=0, current_stock=0)
            result = await db.part_stocks.insert_one(part_stock.dict())
            part_stock = await db.part_stocks.find_one({"_id": result.inserted_id})
        
        new_stock = part_stock['current_stock'] + purchase.quantity
        await db.part_stocks.update_one(
            {"_id": part_stock['_id']},
            {"$set": {"current_stock": new_stock}}
        )
        
        # Update purchase record
        purchase_dict = purchase.dict()
        purchase_dict['item_name'] = item['name']
        await db.purchases.update_one(
            {"_id": ObjectId(purchase_id)},
            {"$set": purchase_dict}
        )
        
        updated = await db.purchases.find_one({"_id": ObjectId(purchase_id)})
        return serialize_doc(updated)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating purchase: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str):
    try:
        # Get existing purchase entry
        existing = await db.purchases.find_one({"_id": ObjectId(purchase_id)})
        if not existing:
            raise HTTPException(status_code=404, detail="Purchase entry not found")
        
        # Reduce part stock (reverse the purchase)
        part_stock = await db.part_stocks.find_one({"part_name": existing['part_name']})
        if part_stock:
            new_stock = part_stock['current_stock'] - existing['quantity']
            await db.part_stocks.update_one(
                {"_id": part_stock['_id']},
                {"$set": {"current_stock": new_stock}}
            )
        
        # Delete purchase record
        await db.purchases.delete_one({"_id": ObjectId(purchase_id)})
        
        return {"message": "Purchase entry deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting purchase: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============= RESET DATABASE =============

@api_router.post("/reset-database")
async def reset_database():
    try:
        await db.items.delete_many({})
        await db.part_stocks.delete_many({})
        await db.production.delete_many({})
        await db.sales.delete_many({})
        await db.purchases.delete_many({})
        return {"message": "Database reset successfully"}
    except Exception as e:
        logger.error(f"Error resetting database: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============= DETAILED REPORTS =============

@api_router.get("/reports/item-details/{item_id}")
async def get_item_details(item_id: str):
    try:
        item = await db.items.find_one({"_id": ObjectId(item_id)})
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        
        # Get production records
        production = await db.production.find({"item_id": item_id}).sort("date", -1).to_list(1000)
        
        # Get sales records
        sales = await db.sales.find({"item_id": item_id}).sort("date", -1).to_list(1000)
        
        return {
            "item": serialize_doc(item),
            "production": [serialize_doc(p) for p in production],
            "sales": [serialize_doc(s) for s in sales]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching item details: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Google Drive Configuration
GOOGLE_DRIVE_FOLDER_ID = "16S2JR4or0ZQYBnZGXzwXgntRrdqZLXEd"
GOOGLE_CREDENTIALS_FILE = ROOT_DIR / "google_credentials.json"

def get_drive_service():
    """Initialize Google Drive service"""
    credentials = service_account.Credentials.from_service_account_file(
        str(GOOGLE_CREDENTIALS_FILE),
        scopes=['https://www.googleapis.com/auth/drive.file']
    )
    return build('drive', 'v3', credentials=credentials)

# ============= EXCEL EXPORT =============

@api_router.get("/export/excel")
async def export_excel():
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Items Stock Sheet
        ws_items = wb.create_sheet("Finished Goods Stock")
        ws_items.append(["Item Name", "Category", "Opening Stock", "Current Stock"])
        for cell in ws_items[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        items = await db.items.find().sort("name", 1).to_list(1000)
        for item in items:
            ws_items.append([item['name'], item['category'], item['opening_stock'], item['current_stock']])
        
        # Parts Stock Sheet
        ws_parts = wb.create_sheet("Parts Stock")
        ws_parts.append(["Part Name", "Opening Stock", "Current Stock"])
        for cell in ws_parts[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        parts = await db.part_stocks.find().sort("part_name", 1).to_list(1000)
        for part in parts:
            ws_parts.append([part['part_name'], part['opening_stock'], part['current_stock']])
        
        # Production Report
        ws_prod = wb.create_sheet("Production Report")
        ws_prod.append(["Date", "Item Name", "Quantity"])
        for cell in ws_prod[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        production = await db.production.find().sort("date", -1).to_list(1000)
        for prod in production:
            ws_prod.append([str(prod['date']), prod['item_name'], prod['quantity']])
        
        # Sales Report
        ws_sales = wb.create_sheet("Sales Report")
        ws_sales.append(["Date", "Item Name", "Quantity", "Party Name"])
        for cell in ws_sales[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        sales = await db.sales.find().sort("date", -1).to_list(1000)
        for sale in sales:
            ws_sales.append([str(sale['date']), sale['item_name'], sale['quantity'], sale['party_name']])
        
        # Purchase Report
        ws_purch = wb.create_sheet("Purchase Report")
        ws_purch.append(["Date", "Item Name", "Part Name", "Quantity"])
        for cell in ws_purch[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        purchases = await db.purchases.find().sort("date", -1).to_list(1000)
        for purch in purchases:
            ws_purch.append([str(purch['date']), purch['item_name'], purch['part_name'], purch['quantity']])
        
        # Auto-adjust columns
        for ws in [ws_items, ws_parts, ws_prod, ws_sales, ws_purch]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=bafna_lights_report_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    except Exception as e:
        logger.error(f"Error exporting excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/email-report")
async def email_report():
    try:
        # Generate Excel file
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Items Stock Sheet
        ws_items = wb.create_sheet("Finished Goods Stock")
        ws_items.append(["Item Name", "Category", "Opening Stock", "Current Stock"])
        for cell in ws_items[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        items = await db.items.find().sort("name", 1).to_list(1000)
        for item in items:
            ws_items.append([item['name'], item['category'], item['opening_stock'], item['current_stock']])
        
        # Parts Stock Sheet
        ws_parts = wb.create_sheet("Parts Stock")
        ws_parts.append(["Part Name", "Opening Stock", "Current Stock"])
        for cell in ws_parts[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        parts = await db.part_stocks.find().sort("part_name", 1).to_list(1000)
        for part in parts:
            ws_parts.append([part['part_name'], part['opening_stock'], part['current_stock']])
        
        # Production Report
        ws_prod = wb.create_sheet("Production Report")
        ws_prod.append(["Date", "Item Name", "Quantity"])
        for cell in ws_prod[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        production = await db.production.find().sort("date", -1).to_list(1000)
        for prod in production:
            ws_prod.append([str(prod['date']), prod['item_name'], prod['quantity']])
        
        # Sales Report
        ws_sales = wb.create_sheet("Sales Report")
        ws_sales.append(["Date", "Item Name", "Quantity", "Party Name"])
        for cell in ws_sales[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        sales = await db.sales.find().sort("date", -1).to_list(1000)
        for sale in sales:
            ws_sales.append([str(sale['date']), sale['item_name'], sale['quantity'], sale['party_name']])
        
        # Purchase Report
        ws_purch = wb.create_sheet("Purchase Report")
        ws_purch.append(["Date", "Item Name", "Part Name", "Quantity"])
        for cell in ws_purch[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        purchases = await db.purchases.find().sort("date", -1).to_list(1000)
        for purch in purchases:
            ws_purch.append([str(purch['date']), purch['item_name'], purch['part_name'], purch['quantity']])
        
        # Auto-adjust columns
        for ws in [ws_items, ws_parts, ws_prod, ws_sales, ws_purch]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Send Email
        filename = f"bafna_lights_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        gmail_user = os.environ.get('GMAIL_USER')
        gmail_password = os.environ.get('GMAIL_APP_PASSWORD')
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = gmail_user
        msg['To'] = gmail_user
        msg['Subject'] = f"Bafna Light's Stock Report - {datetime.now().strftime('%d %B %Y')}"
        
        body = f"""
Dear Team,

Please find attached the daily stock report for Bafna Light's.

Report Date: {datetime.now().strftime('%d %B %Y')}
Generated: {datetime.now().strftime('%I:%M %p')}

This report includes:
- Finished Goods Stock
- Parts Stock
- Production Report
- Sales Report
- Purchase Report

Best regards,
Bafna Light's Stock Management System
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach Excel file
        attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        attachment.set_payload(excel_file.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
        msg.attach(attachment)
        
        # Send email
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(gmail_user, gmail_password)
        text = msg.as_string()
        server.sendmail(gmail_user, gmail_user, text)
        server.quit()
        
        return {
            "success": True,
            "message": f"Report emailed to {gmail_user} successfully!",
            "filename": filename
        }
    except Exception as e:
        logger.error(f"Error emailing report: {e}")
        raise HTTPException(status_code=500, detail=str(e))

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
